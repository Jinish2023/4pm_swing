"""
Styled Excel portfolio tracker with a real state machine:
    PENDING ENTRY  -->  OPEN  -->  EXITED
Two sheets: "Dashboard" (KPI summary cards) and "Positions" (the full log).
"""
import os
from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
POSITIONS_SHEET = "Positions"
DASHBOARD_SHEET = "Dashboard"

# ── Palette (vibrant, per your request) ──
HEADER_FILL = "3730A3"       # deep indigo
HEADER_FONT_COLOR = "FFFFFF"
PENDING_FILL = "FDE68A"      # amber
PENDING_FONT = "78350F"
OPEN_FILL = "10B981"         # emerald
OPEN_FONT = "FFFFFF"
EXIT_WIN_FILL = "0EA5E9"     # sky blue
EXIT_WIN_FONT = "FFFFFF"
EXIT_LOSS_FILL = "EF4444"    # red
EXIT_LOSS_FONT = "FFFFFF"
PNL_POS_FILL = "D1FAE5"
PNL_POS_FONT = "065F46"
PNL_NEG_FILL = "FEE2E2"
PNL_NEG_FONT = "991B1B"
ZEBRA_FILL = "F8FAFC"
BORDER_COLOR = "CBD5E1"

POSITIONS_HEADERS = [
    "Serial No", "Signal Date", "Stock", "Entry Type", "Status",
    "Signal Close (Rs)", "ATR @ Signal", "Entry Date", "Entry Price (Rs)",
    "Initial Stop (3xATR)", "Current Price (Rs)", "Highest Close Since Entry",
    "Current Trailing Stop", "20 SMA (Latest)", "P&L (Rs)", "P&L %",
    "Exit Date", "Exit Price (Rs)", "Exit Reason", "Days Held", "Last Updated (IST)",
]
COL_WIDTHS = [9, 13, 12, 14, 14, 13, 11, 13, 13, 15, 13, 18, 16, 13, 11, 9, 13, 13, 30, 10, 18]

_thin = Side(style="thin", color=BORDER_COLOR)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _load_or_create(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet(DASHBOARD_SHEET)
    ws = wb.create_sheet(POSITIONS_SHEET)
    _style_positions_header(ws)
    return wb


def _style_positions_header(ws):
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, title in enumerate(POSITIONS_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, _border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(POSITIONS_HEADERS))}1"
    ws.sheet_view.showGridLines = False


def _status_fill_font(status, pnl_pct):
    if status == "PENDING ENTRY":
        return PENDING_FILL, PENDING_FONT
    if status == "OPEN":
        return OPEN_FILL, OPEN_FONT
    if status == "EXITED":
        if pnl_pct is not None and pnl_pct >= 0:
            return EXIT_WIN_FILL, EXIT_WIN_FONT
        return EXIT_LOSS_FILL, EXIT_LOSS_FONT
    return "FFFFFF", "000000"


def _style_row(ws, r, zebra):
    body_font = Font(name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center")
    zebra_fill = PatternFill("solid", fgColor=ZEBRA_FILL) if zebra else None
    for c in range(1, len(POSITIONS_HEADERS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = center
        cell.border = _border
        if zebra_fill:
            cell.fill = zebra_fill

    for col_letter in ["F", "I", "J", "K", "L", "M", "N", "O"]:
        ws[f"{col_letter}{r}"].number_format = "#,##0.00"
    ws[f"P{r}"].number_format = "+0.00%;-0.00%"

    status = ws[f"E{r}"].value
    pnl_pct_raw = ws[f"P{r}"].value
    pnl_pct = pnl_pct_raw if isinstance(pnl_pct_raw, (int, float)) else None
    fill_hex, font_hex = _status_fill_font(status, pnl_pct)
    status_cell = ws[f"E{r}"]
    status_cell.fill = PatternFill("solid", fgColor=fill_hex)
    status_cell.font = Font(name="Calibri", size=10, bold=True, color=font_hex)

    pnl_cell = ws[f"P{r}"]
    if pnl_pct is not None:
        if pnl_pct >= 0:
            pnl_cell.fill = PatternFill("solid", fgColor=PNL_POS_FILL)
            pnl_cell.font = Font(name="Calibri", size=10, bold=True, color=PNL_POS_FONT)
        else:
            pnl_cell.fill = PatternFill("solid", fgColor=PNL_NEG_FILL)
            pnl_cell.font = Font(name="Calibri", size=10, bold=True, color=PNL_NEG_FONT)


# NOTE: column letters below map 1:1 to POSITIONS_HEADERS order.
COL = {name: get_column_letter(i + 1) for i, name in enumerate([
    "serial", "signal_date", "stock", "entry_type", "status", "signal_close",
    "atr_signal", "entry_date", "entry_price", "initial_stop", "current_price",
    "highest_close", "trailing_stop", "sma20", "pnl_rs", "pnl_pct",
    "exit_date", "exit_price", "exit_reason", "days_held", "last_updated",
])}


def get_tracked_rows_full(path):
    """Returns {row_idx: {symbol, status, signal_date, atr_signal, entry_date,
    entry_price, highest_close, trailing_stop}} for every PENDING/OPEN row —
    everything needed to drive the state-machine update each run."""
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path)
    if POSITIONS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[POSITIONS_SHEET]
    out = {}
    for r in range(2, ws.max_row + 1):
        status = ws[f"{COL['status']}{r}"].value
        symbol = ws[f"{COL['stock']}{r}"].value
        if not symbol or status not in ("PENDING ENTRY", "OPEN"):
            continue
        signal_date_str = ws[f"{COL['signal_date']}{r}"].value
        entry_date_str = ws[f"{COL['entry_date']}{r}"].value
        out[r] = {
            "symbol": symbol,
            "status": status,
            "signal_date": datetime.strptime(signal_date_str, "%d/%m/%Y") if signal_date_str else None,
            "atr_signal": ws[f"{COL['atr_signal']}{r}"].value,
            "entry_date": datetime.strptime(entry_date_str, "%d/%m/%Y") if entry_date_str else None,
            "entry_price": ws[f"{COL['entry_price']}{r}"].value or None,
            "highest_close": ws[f"{COL['highest_close']}{r}"].value or None,
            "trailing_stop": ws[f"{COL['trailing_stop']}{r}"].value or None,
        }
    return out


def get_tracked_symbols(path):
    """Returns {symbol: [row_idx, ...]} for rows currently PENDING or OPEN
    (used to dedup new signals and to know what to update each run)."""
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path)
    if POSITIONS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[POSITIONS_SHEET]
    tracked = {}
    for r in range(2, ws.max_row + 1):
        status = ws[f"{COL['status']}{r}"].value
        symbol = ws[f"{COL['stock']}{r}"].value
        if symbol and status in ("PENDING ENTRY", "OPEN"):
            tracked.setdefault(symbol, []).append(r)
    return tracked


def append_pending_signals(path, signals):
    """Appends new PENDING ENTRY rows. Caller is responsible for having
    already excluded symbols with an existing open/pending row (dedup)."""
    wb = _load_or_create(path)
    ws = wb[POSITIONS_SHEET] if POSITIONS_SHEET in wb.sheetnames else wb.create_sheet(POSITIONS_SHEET)
    if ws.max_row == 1 and ws["A1"].value != POSITIONS_HEADERS[0]:
        _style_positions_header(ws)

    next_serial = ws.max_row - 1  # header is row 1
    now_str = datetime.now(IST).strftime("%d/%m/%Y %I:%M %p IST")

    for sig in signals:
        next_serial += 1
        r = next_serial + 1
        initial_stop = round(sig["signal_close"] - 3.0 * sig["atr_at_signal"], 2)
        row_values = {
            "serial": next_serial, "signal_date": sig["signal_date"].strftime("%d/%m/%Y"),
            "stock": sig["symbol"], "entry_type": sig["entry_type"], "status": "PENDING ENTRY",
            "signal_close": round(sig["signal_close"], 2), "atr_signal": round(sig["atr_at_signal"], 2),
            "entry_date": "", "entry_price": "", "initial_stop": initial_stop,
            "current_price": round(sig["signal_close"], 2), "highest_close": "", "trailing_stop": "",
            "sma20": "", "pnl_rs": "", "pnl_pct": "", "exit_date": "", "exit_price": "",
            "exit_reason": "", "days_held": 0, "last_updated": now_str,
        }
        for key, val in row_values.items():
            ws[f"{COL[key]}{r}"] = val
        _style_row(ws, r, zebra=(next_serial % 2 == 0))

    wb.save(path)
    return len(signals)


def update_tracked_rows(path, updates):
    """
    updates: {row_idx: {...fields to overwrite...}} — see 4pm_swing.py for
    exactly which fields get passed for PENDING->OPEN confirmation vs OPEN
    daily updates vs OPEN->EXITED transitions.
    """
    if not updates:
        return
    wb = openpyxl.load_workbook(path)
    ws = wb[POSITIONS_SHEET]
    now_str = datetime.now(IST).strftime("%d/%m/%Y %I:%M %p IST")

    for r, fields in updates.items():
        for key, val in fields.items():
            ws[f"{COL[key]}{r}"] = val
        ws[f"{COL['last_updated']}{r}"] = now_str
        _style_row(ws, r, zebra=((r - 1) % 2 == 0))

    wb.save(path)


def rebuild_dashboard(path):
    wb = openpyxl.load_workbook(path)
    if POSITIONS_SHEET not in wb.sheetnames:
        return
    ws = wb[POSITIONS_SHEET]

    pending = open_ = exited = wins = losses = 0
    pnl_pcts = []
    for r in range(2, ws.max_row + 1):
        status = ws[f"{COL['status']}{r}"].value
        if status == "PENDING ENTRY":
            pending += 1
        elif status == "OPEN":
            open_ += 1
        elif status == "EXITED":
            exited += 1
            pnl = ws[f"{COL['pnl_pct']}{r}"].value
            if isinstance(pnl, (int, float)):
                pnl_pcts.append(pnl)
                if pnl >= 0:
                    wins += 1
                else:
                    losses += 1

    win_rate = round(100 * wins / exited, 1) if exited else 0.0
    avg_pnl = round(100 * (sum(pnl_pcts) / len(pnl_pcts)), 2) if pnl_pcts else 0.0
    best = round(100 * max(pnl_pcts), 2) if pnl_pcts else 0.0
    worst = round(100 * min(pnl_pcts), 2) if pnl_pcts else 0.0

    if DASHBOARD_SHEET in wb.sheetnames:
        wb.remove(wb[DASHBOARD_SHEET])
    dash = wb.create_sheet(DASHBOARD_SHEET, 0)
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 3
    for col in "BCDEF":
        dash.column_dimensions[col].width = 20

    title_font = Font(name="Calibri", size=18, bold=True, color="1E1B4B")
    dash["B2"] = "4PM Swing — Portfolio Dashboard"
    dash["B2"].font = title_font
    dash["B3"] = f"Last updated: {datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}"
    dash["B3"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

    cards = [
        ("PENDING", pending, PENDING_FILL, PENDING_FONT),
        ("OPEN POSITIONS", open_, OPEN_FILL, OPEN_FONT),
        ("EXITED (TOTAL)", exited, "6366F1", "FFFFFF"),
        ("WIN RATE", f"{win_rate}%", "0EA5E9" if win_rate >= 50 else EXIT_LOSS_FILL, "FFFFFF"),
        ("AVG P&L (EXITED)", f"{avg_pnl:+.2f}%", PNL_POS_FILL if avg_pnl >= 0 else PNL_NEG_FILL,
         PNL_POS_FONT if avg_pnl >= 0 else PNL_NEG_FONT),
        ("BEST TRADE", f"{best:+.2f}%", PNL_POS_FILL, PNL_POS_FONT),
        ("WORST TRADE", f"{worst:+.2f}%", PNL_NEG_FILL, PNL_NEG_FONT),
    ]

    row, col = 5, 2
    for i, (label, value, fill_hex, font_hex) in enumerate(cards):
        c1 = dash.cell(row=row, column=col)
        c2 = dash.cell(row=row + 1, column=col)
        dash.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        dash.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        c1.value, c2.value = label, value
        c1.font = Font(name="Calibri", size=10, bold=True, color=font_hex)
        c2.font = Font(name="Calibri", size=20, bold=True, color=font_hex)
        c1.alignment = c2.alignment = Alignment(horizontal="center", vertical="center")
        for cc in (c1, c2):
            cc.fill = PatternFill("solid", fgColor=fill_hex)
        dash.row_dimensions[row].height = 22
        dash.row_dimensions[row + 1].height = 32
        col += 2
        if col > 6:
            col = 2
            row += 3

    wb.save(path)
